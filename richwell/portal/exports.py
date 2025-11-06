"""
Export utilities for generating PDF and CSV/Excel reports.
"""

import csv
from io import BytesIO
from datetime import datetime
from django.http import HttpResponse
from reportlab.lib import colors
from reportlab.lib.pagesizes import letter, A4
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, PageBreak
from reportlab.lib.enums import TA_CENTER, TA_LEFT, TA_RIGHT
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill

from .models import Student, StudentSubject, Section, Grade, Term


class PDFExporter:
    """Handles PDF export functionality."""

    @staticmethod
    def generate_cor(student, term):
        """
        Generate Certificate of Registration (COR) for a student.

        Args:
            student: Student instance
            term: Term instance

        Returns:
            HttpResponse with PDF content
        """
        # Create response
        response = HttpResponse(content_type='application/pdf')
        response['Content-Disposition'] = f'attachment; filename="COR_{student.user.username}_{term.name}.pdf"'

        # Create PDF
        buffer = BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=letter)
        elements = []

        # Styles
        styles = getSampleStyleSheet()
        title_style = ParagraphStyle(
            'CustomTitle',
            parent=styles['Heading1'],
            fontSize=18,
            textColor=colors.HexColor('#1f2937'),
            spaceAfter=30,
            alignment=TA_CENTER
        )

        heading_style = ParagraphStyle(
            'CustomHeading',
            parent=styles['Heading2'],
            fontSize=14,
            textColor=colors.HexColor('#374151'),
            spaceAfter=12
        )

        # Header
        title = Paragraph("RICHWELL SCHOOL PORTAL", title_style)
        elements.append(title)

        subtitle = Paragraph("Certificate of Registration", heading_style)
        elements.append(subtitle)
        elements.append(Spacer(1, 0.3 * inch))

        # Student Information
        student_data = [
            ["Student Information", ""],
            ["Student ID:", student.user.username],
            ["Name:", f"{student.user.first_name} {student.user.last_name}"],
            ["Program:", student.program.name],
            ["Curriculum:", student.curriculum.version],
            ["Term:", term.name],
        ]

        student_table = Table(student_data, colWidths=[2*inch, 4*inch])
        student_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (1, 0), colors.HexColor('#3b82f6')),
            ('TEXTCOLOR', (0, 0), (1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 12),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 1), (-1, -1), colors.white),
            ('GRID', (0, 0), (-1, -1), 1, colors.grey),
        ]))
        elements.append(student_table)
        elements.append(Spacer(1, 0.3 * inch))

        # Enrolled Subjects
        enrollments = StudentSubject.objects.filter(
            student=student,
            term=term,
            status='enrolled'
        ).select_related('subject', 'section', 'section__professor')

        if enrollments.exists():
            subjects_heading = Paragraph("Enrolled Subjects", heading_style)
            elements.append(subjects_heading)
            elements.append(Spacer(1, 0.2 * inch))

            # Subjects table
            subject_data = [["Code", "Subject Title", "Units", "Section", "Professor"]]
            total_units = 0

            for enrollment in enrollments:
                subject = enrollment.subject
                section = enrollment.section
                subject_data.append([
                    subject.code,
                    subject.title,
                    str(subject.units),
                    section.section_code,
                    f"{section.professor.first_name} {section.professor.last_name}"
                ])
                total_units += subject.units

            subject_data.append(["", "", "", "Total Units:", str(total_units)])

            subjects_table = Table(subject_data, colWidths=[1*inch, 2.5*inch, 0.7*inch, 1*inch, 1.5*inch])
            subjects_table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#3b82f6')),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 10),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                ('BACKGROUND', (0, 1), (-1, -2), colors.white),
                ('GRID', (0, 0), (-1, -1), 1, colors.grey),
                ('BACKGROUND', (0, -1), (-1, -1), colors.HexColor('#e5e7eb')),
                ('FONTNAME', (0, -1), (-1, -1), 'Helvetica-Bold'),
            ]))
            elements.append(subjects_table)
        else:
            no_enrollment = Paragraph("No subjects enrolled for this term.", styles['Normal'])
            elements.append(no_enrollment)

        elements.append(Spacer(1, 0.5 * inch))

        # Footer
        footer_text = f"Generated on: {datetime.now().strftime('%B %d, %Y at %I:%M %p')}"
        footer = Paragraph(footer_text, styles['Normal'])
        elements.append(footer)

        # Build PDF
        doc.build(elements)
        pdf = buffer.getvalue()
        buffer.close()
        response.write(pdf)

        return response

    @staticmethod
    def generate_section_roster(section):
        """
        Generate PDF roster for a section.

        Args:
            section: Section instance

        Returns:
            HttpResponse with PDF content
        """
        response = HttpResponse(content_type='application/pdf')
        response['Content-Disposition'] = f'attachment; filename="Roster_{section.section_code}.pdf"'

        buffer = BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=letter)
        elements = []

        styles = getSampleStyleSheet()
        title_style = ParagraphStyle(
            'CustomTitle',
            parent=styles['Heading1'],
            fontSize=16,
            textColor=colors.HexColor('#1f2937'),
            spaceAfter=20,
            alignment=TA_CENTER
        )

        # Header
        title = Paragraph(f"Class Roster - {section.section_code}", title_style)
        elements.append(title)
        elements.append(Spacer(1, 0.2 * inch))

        # Section info
        section_data = [
            ["Subject:", f"{section.subject.code} - {section.subject.title}"],
            ["Term:", section.term.name],
            ["Professor:", f"{section.professor.first_name} {section.professor.last_name}"],
            ["Capacity:", f"{section.studentsubject_set.count()} / {section.capacity}"],
        ]

        section_table = Table(section_data, colWidths=[1.5*inch, 5*inch])
        section_table.setStyle(TableStyle([
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('FONTNAME', (0, 0), (0, -1), 'Helvetica-Bold'),
            ('GRID', (0, 0), (-1, -1), 1, colors.grey),
        ]))
        elements.append(section_table)
        elements.append(Spacer(1, 0.3 * inch))

        # Student roster
        enrollments = section.studentsubject_set.filter(status='enrolled').select_related('student__user', 'student__program')

        roster_data = [["No.", "Student ID", "Name", "Program"]]
        for i, enrollment in enumerate(enrollments, 1):
            student = enrollment.student
            roster_data.append([
                str(i),
                student.user.username,
                f"{student.user.last_name}, {student.user.first_name}",
                student.program.name
            ])

        roster_table = Table(roster_data, colWidths=[0.5*inch, 1.5*inch, 3*inch, 2*inch])
        roster_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#3b82f6')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('GRID', (0, 0), (-1, -1), 1, colors.grey),
        ]))
        elements.append(roster_table)

        # Build PDF
        doc.build(elements)
        pdf = buffer.getvalue()
        buffer.close()
        response.write(pdf)

        return response


class CSVExporter:
    """Handles CSV export functionality."""

    @staticmethod
    def export_grades(section):
        """
        Export grades for a section to CSV.

        Args:
            section: Section instance

        Returns:
            HttpResponse with CSV content
        """
        response = HttpResponse(content_type='text/csv')
        response['Content-Disposition'] = f'attachment; filename="Grades_{section.section_code}.csv"'

        writer = csv.writer(response)

        # Header
        writer.writerow(['Student ID', 'Last Name', 'First Name', 'Numeric Grade', 'Letter Grade', 'Status'])

        # Get enrollments and grades
        enrollments = section.studentsubject_set.filter(status__in=['enrolled', 'completed']).select_related('student__user')

        for enrollment in enrollments:
            student = enrollment.student
            try:
                grade = Grade.objects.get(student_subject=enrollment)
                numeric = grade.numeric_grade if grade.numeric_grade else ''
                letter = grade.letter_grade if grade.letter_grade else ''
            except Grade.DoesNotExist:
                numeric = ''
                letter = ''

            writer.writerow([
                student.user.username,
                student.user.last_name,
                student.user.first_name,
                numeric,
                letter,
                enrollment.status
            ])

        return response

    @staticmethod
    def export_enrollment_report(term):
        """
        Export enrollment statistics for a term to CSV.

        Args:
            term: Term instance

        Returns:
            HttpResponse with CSV content
        """
        response = HttpResponse(content_type='text/csv')
        response['Content-Disposition'] = f'attachment; filename="Enrollment_Report_{term.name}.csv"'

        writer = csv.writer(response)

        # Header
        writer.writerow(['Section Code', 'Subject', 'Professor', 'Enrolled', 'Capacity', 'Status'])

        # Get sections for term
        sections = Section.objects.filter(term=term).select_related('subject', 'professor')

        for section in sections:
            enrolled_count = section.studentsubject_set.filter(status='enrolled').count()
            writer.writerow([
                section.section_code,
                f"{section.subject.code} - {section.subject.title}",
                f"{section.professor.first_name} {section.professor.last_name}",
                enrolled_count,
                section.capacity,
                section.status
            ])

        return response


class ExcelExporter:
    """Handles Excel export functionality using openpyxl."""

    @staticmethod
    def export_grades_excel(section):
        """
        Export grades for a section to Excel.

        Args:
            section: Section instance

        Returns:
            HttpResponse with Excel content
        """
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = f'attachment; filename="Grades_{section.section_code}.xlsx"'

        # Create workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "Grades"

        # Header style
        header_fill = PatternFill(start_color="3B82F6", end_color="3B82F6", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")
        header_alignment = Alignment(horizontal="center", vertical="center")

        # Write header
        headers = ['Student ID', 'Last Name', 'First Name', 'Numeric Grade', 'Letter Grade', 'Status']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = header_alignment

        # Write data
        enrollments = section.studentsubject_set.filter(status__in=['enrolled', 'completed']).select_related('student__user')

        for row, enrollment in enumerate(enrollments, 2):
            student = enrollment.student
            try:
                grade = Grade.objects.get(student_subject=enrollment)
                numeric = grade.numeric_grade if grade.numeric_grade else ''
                letter = grade.letter_grade if grade.letter_grade else ''
            except Grade.DoesNotExist:
                numeric = ''
                letter = ''

            ws.cell(row=row, column=1, value=student.user.username)
            ws.cell(row=row, column=2, value=student.user.last_name)
            ws.cell(row=row, column=3, value=student.user.first_name)
            ws.cell(row=row, column=4, value=numeric)
            ws.cell(row=row, column=5, value=letter)
            ws.cell(row=row, column=6, value=enrollment.status)

        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(cell.value)
                except:
                    pass
            adjusted_width = (max_length + 2)
            ws.column_dimensions[column_letter].width = adjusted_width

        # Save to response
        wb.save(response)
        return response
